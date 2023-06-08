import os
import json
import pandas as pd
import numpy
import datetime
import config

from openpyxl import load_workbook

from components.ExcelConnector.custom_forms.extract_form import ExcelExtractForm
from components.ExcelConnector.custom_forms.init_form import ExcelInitForm
from components.ExcelConnector.custom_forms.metadata_form import ExcelMetadataForm
from model_srv.BaseComponent.BaseConnectorComponent import BaseConnectorComponent
from model_srv.mongodb.CObjectService import BackendCObject


class ExcelConnection(BaseConnectorComponent):

    def __init__(self, bk_object: BackendCObject):
        self.bk_object = bk_object
        if bk_object.properties['filename']['value']:
            self.filename = os.path.basename(bk_object.properties['filename']['value'])
        else:
            self.filename = ''
        if bk_object.properties['filename']['value']:
            self.base_dir = os.path.dirname(bk_object.properties['filename']['value'])
        else:
            self.base_dir = ''

        self.queries = dict()
        self.data = dict()

    def get_projection(self, previous=None):
        self.build_queries(self.bk_object.properties['q_data']['value'])
        for filename, sheet_columns in self.queries.items():
            for sheet, columns in sheet_columns.items():
                self.extract((filename, sheet), columns)
        result = list()
        for k, v in self.data.items():
            v['table_name'] = k
            result.append(v)
        return result

    def get_possible_values(self, _property) -> list:
        bk_category = self.bk_object.get_bk_category()
        for t in bk_category.template:
            if t['self_name'] == _property:
                return t['possible_values']
        return list()

    def init_form(self, model_form):
        form = ExcelInitForm(model_form, self)
        form.show()

    def get_metadata(self, model_form):
        metadata_excel = list()
        filepath = self._get_filepath(self.filename)

        workbook = load_workbook(filepath, read_only=True)
        data = {}
        for sheet in workbook.worksheets:
            for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                data[sheet.title] = value

        metadata_excel.append(self.transform_mdt(data, self.filename))

        form = ExcelMetadataForm(model_form, self, metadata_excel)
        form.show()

    def build_queries(self, _q_data):
        if isinstance(_q_data, str):
            _q_data = json.loads(_q_data)
        r = dict()
        for filename, sheets in _q_data.items():
            r[filename] = dict()
            if sheets is None:
                continue
            for sheet, columns in sheets.items():
                if not isinstance(columns, dict):
                    continue
                columns = {k: v for k, v in columns.items() if isinstance(v, dict)}
                r[filename][sheet] = columns

        self.queries = r
        return r

    def extract(self, fn_sheet, columns):
        filepath = self._get_filepath(fn_sheet[0])
        return_data = self.read_and_processing(filepath, sheet_name=fn_sheet[1], usecols=list(columns.keys()), index_col=None)
        self.data[fn_sheet[1]] = return_data

    def extract_all(self, model_form):
        self.build_queries(self.bk_object.properties['q_data']['value'])
        for filename, sheet_columns in self.queries.items():
            for sheet, columns in sheet_columns.items():
                self.extract((filename, sheet), columns)

        form = ExcelExtractForm(model_form, self, self.data)
        form.show()

    def transform_mdt(self, _data, _fn):
        r = {
            'ИмяЭлемента': _fn,
            'ТипЭлемента': 'ИмяФайла',
            'row': []
        }
        for sheet, columns in _data.items():
            rows = list()
            for col in columns:
                if isinstance(col, datetime.datetime):
                    col = col.strftime('%Y-%m-%d %H:%M:%S')
                row_dct = {
                    'ИмяЭлемента': col,
                    'ТипЭлемента': 'ИмяКолонки'
                }
                rows.append(row_dct)

            sheet_dct = {
                'ИмяЭлемента': sheet,
                'ТипЭлемента': 'ИмяЛиста',
                'row': rows
            }
            r['row'].append(sheet_dct)
        return r

    def _get_filepath(self, fn):
        return os.path.join(self.base_dir, fn)

    def _get_tmp_filepath(self, fn):
        now = datetime.datetime.now()
        # now_datetime = now.strftime('%d_%m_%Y__%H_%M')
        now_datetime = now.strftime('%d_%m_%Y')
        return os.path.join(self.base_dir, f'{now_datetime}_{fn}')

    def load_tbl(self, tbl, table_name, field_map):
        if len(self.filename) > 1:
            return

        filepath = self._get_tmp_filepath(self.filename[0])

        # filepath = os.path.join(self.base_dir, self.filename[0])

        df = pd.DataFrame(tbl['data'], columns=tbl['columns'])
        for col in [col for col in df.columns if col not in field_map.keys()]:
            df.drop(col, axis=1, inplace=True)
        df.rename(columns=field_map, inplace=True)
        if not df.empty:
            df.to_excel(filepath, sheet_name=table_name, index=False, encoding='utf-8')

    def read_and_processing(self, filepath, sheet_name, usecols=None, index_col=None):
        date_col_called = False
        if usecols:
            usecols = list(filter(lambda x: x is not None and x != 'null', usecols))

            for col_id in range(len(usecols)):
                try:
                    usecols[col_id] = datetime.datetime.strptime(usecols[col_id], '%Y-%m-%d %H:%M:%S')
                    date_col_called = True
                except ValueError:
                    continue
        if date_col_called:
            df = pd.read_excel(filepath, sheet_name=sheet_name, index_col=index_col)[usecols]
        else:
            df = pd.read_excel(filepath, sheet_name=sheet_name, index_col=index_col, usecols=usecols)
        excel_file = load_workbook(filename=filepath)
        sheet = excel_file[sheet_name]

        for r in sheet.merged_cells.ranges:
            cl, rl, cr, rr = r.bounds  # границы объединенной области
            rl -= 2
            rr -= 1
            cl -= 1
            base_value = df.iloc[rl, cl]
            df.iloc[rl:rr, cl:cr] = base_value

        for col, t in zip(df.dtypes.index, df.dtypes):
            # if t in ['datetime64[ns]', 'timedelta64[ns]']:
            if any(i for i in df[col] if isinstance(i, datetime.datetime) or isinstance(i, datetime.timedelta)):
                df[col] = df[col].astype('str')

        df.replace(numpy.nan, 'NaN', inplace=True)

        return_data = json.loads(df.to_json(orient='split'))

        return return_data
